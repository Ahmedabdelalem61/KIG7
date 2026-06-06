import base64
import io

from openpyxl import Workbook, load_workbook

from odoo.exceptions import UserError
from odoo.tests.common import TransactionCase, tagged


@tagged("post_install", "-at_install")
class TestXlsxIo(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Wizard = cls.env["hr.uae.xlsx.import.wizard"]
        cls.Template = cls.env["hr.uae.xlsx.template"].with_context(active_test=False)
        # Predefined templates are seeded by the data file at install.
        cls.tpl_emp = cls.Template.search([("code", "=", "employees")], limit=1)
        cls.tpl_payroll = cls.Template.search(
            [("code", "=", "payroll_dashboard")], limit=1
        )
        cls.emp = cls.env["hr.employee"].create(
            {"name": "XLSX Base Employee", "work_phone": "111"}
        )

    # ------------------------------------------------------------------ utils

    def _make_file(self, headers, rows, tech_row=None, title="Data"):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(headers)
        if tech_row is not None:
            ws.append(tech_row)
        for row in rows:
            ws.append(row)
        buffer = io.BytesIO()
        wb.save(buffer)
        return base64.b64encode(buffer.getvalue())

    def _new_wizard(self, template, file_b64=None):
        return self.Wizard.create(
            {
                "template_id": template.id,
                "import_file": file_b64,
                "import_filename": "test.xlsx" if file_b64 else False,
            }
        )

    def _attachment_from_action(self, action):
        self.assertEqual(action["type"], "ir.actions.act_url")
        att_id = int(action["url"].split("/web/content/")[1].split("?")[0])
        att = self.env["ir.attachment"].browse(att_id)
        return load_workbook(io.BytesIO(base64.b64decode(att.datas)))

    # ------------------------------------------------------------------ seed

    def test_predefined_templates_installed(self):
        for code in [
            "employees",
            "contracts",
            "flights",
            "documents",
            "salary_adjustments",
            "terminations",
            "payroll_dashboard",
        ]:
            tpl = self.Template.search([("code", "=", code)], limit=1)
            self.assertTrue(tpl, "Template %s should be seeded" % code)
            self.assertTrue(tpl.line_ids, "Template %s should have lines" % code)

        self.assertFalse(self.tpl_payroll.allow_import)
        self.assertFalse(self.tpl_payroll.import_line_ids())
        # employees has a match key (passport)
        self.assertIn("passport_id", self.tpl_emp.match_key_line_ids().mapped("name"))

    # ------------------------------------------------------------------ payroll

    def test_payroll_basic_net_fields(self):
        model = self.env["hr.uae.payroll.dashboard"]
        self.assertIn("basic_amount", model._fields)
        self.assertIn("net_amount", model._fields)
        rows = model.search([], limit=20)
        for row in rows:
            # Net mirrors Total to Pay.
            self.assertAlmostEqual(row.net_amount, row.total_to_pay, places=2)

    def test_payroll_import_blocked(self):
        wiz = self._new_wizard(self.tpl_payroll)
        with self.assertRaises(UserError):
            wiz.action_import()

    # ------------------------------------------------------------------ export

    def test_export_has_sheets_and_metadata(self):
        wiz = self._new_wizard(self.tpl_emp)
        action = wiz.action_export_data()
        wb = self._attachment_from_action(action)
        self.assertIn("Data", wb.sheetnames)
        self.assertIn("Instructions", wb.sheetnames)
        ws = wb["Data"]
        rows = list(ws.iter_rows(values_only=True))
        # Row 1 = labels, Row 2 = technical names, first column is the id.
        self.assertEqual(str(rows[0][0]), "ID")
        self.assertEqual(str(rows[1][0]), "id")
        self.assertIn("name", [str(c) for c in rows[1]])

    def test_export_payroll_has_basic_and_net(self):
        wiz = self._new_wizard(self.tpl_payroll)
        wb = self._attachment_from_action(wiz.action_export_data())
        tech = [str(c) for c in list(wb["Data"].iter_rows(values_only=True))[1]]
        self.assertIn("basic_amount", tech)
        self.assertIn("net_amount", tech)

    # ------------------------------------------------------------------ import

    def test_import_create(self):
        before = self.env["hr.employee"].search_count([("name", "=", "Imported A")])
        file_b64 = self._make_file(
            ["name", "work_email", "gender"],
            [["Imported A", "a@example.com", "Male"]],
        )
        wiz = self._new_wizard(self.tpl_emp, file_b64)
        wiz.action_import()
        self.assertEqual(wiz.error_count, 0)
        self.assertEqual(wiz.create_count, 1)
        emp = self.env["hr.employee"].search([("name", "=", "Imported A")])
        self.assertEqual(len(emp), before + 1)
        self.assertEqual(emp[0].work_email, "a@example.com")
        self.assertEqual(emp[0].gender, "male")

    def test_import_update_by_id(self):
        file_b64 = self._make_file(
            ["id", "work_phone"],
            [[self.emp.id, "999-updated"]],
        )
        wiz = self._new_wizard(self.tpl_emp, file_b64)
        wiz.action_import()
        self.assertEqual(wiz.error_count, 0)
        self.assertEqual(wiz.update_count, 1)
        self.assertEqual(self.emp.work_phone, "999-updated")

    def test_import_update_by_match_key(self):
        target = self.env["hr.employee"].create(
            {"name": "Match Key Emp", "passport_id": "PPMATCH123"}
        )
        file_b64 = self._make_file(
            ["passport_id", "work_phone"],
            [["PPMATCH123", "match-updated"]],
        )
        wiz = self._new_wizard(self.tpl_emp, file_b64)
        wiz.action_import()
        self.assertEqual(wiz.error_count, 0)
        self.assertEqual(wiz.update_count, 1)
        self.assertEqual(target.work_phone, "match-updated")

    def test_import_invalid_selection_blocks_all(self):
        file_b64 = self._make_file(
            ["name", "gender"],
            [
                ["Good Row XLSX", "Male"],
                ["Bad Row XLSX", "not-a-gender"],
            ],
        )
        wiz = self._new_wizard(self.tpl_emp, file_b64)
        wiz.action_import()
        self.assertEqual(wiz.error_count, 1)
        self.assertTrue(wiz.error_file)
        # All-or-nothing: the good row must NOT have been persisted.
        self.assertFalse(
            self.env["hr.employee"].search([("name", "=", "Good Row XLSX")])
        )

    def test_validate_does_not_persist(self):
        file_b64 = self._make_file(
            ["name"],
            [["Validate Only Emp"]],
        )
        wiz = self._new_wizard(self.tpl_emp, file_b64)
        wiz.action_validate()
        self.assertEqual(wiz.error_count, 0)
        self.assertEqual(wiz.create_count, 1)
        # Nothing saved in validate mode.
        self.assertFalse(
            self.env["hr.employee"].search([("name", "=", "Validate Only Emp")])
        )

    def test_import_two_row_roundtrip(self):
        # Export then re-import the exact file (two-row format) to update.
        wiz = self._new_wizard(self.tpl_emp)
        wb = self._attachment_from_action(wiz.action_export_data())
        ws = wb["Data"]
        rows = list(ws.iter_rows(values_only=True))
        labels, tech = list(rows[0]), list(rows[1])
        # find our base employee row and tweak the roster column
        id_idx = tech.index("id")
        data_rows = []
        for r in rows[2:]:
            r = list(r)
            if r[id_idx] == self.emp.id:
                if "roster" in tech:
                    r[tech.index("roster")] = "RT-ROUNDTRIP"
            data_rows.append(r)
        file_b64 = self._make_file(labels, data_rows, tech_row=tech)
        wiz2 = self._new_wizard(self.tpl_emp, file_b64)
        wiz2.action_import()
        self.assertEqual(wiz2.error_count, 0)
        self.emp.invalidate_recordset()
        self.assertEqual(self.emp.roster, "RT-ROUNDTRIP")

    def test_create_custom_template(self):
        model = self.env["ir.model"].search([("model", "=", "res.partner")], limit=1)
        field_name = self.env["ir.model.fields"].search(
            [("model", "=", "res.partner"), ("name", "=", "name")], limit=1
        )
        tpl = self.Template.create(
            {
                "name": "Partners (test)",
                "code": "test_partners",
                "model_id": model.id,
                "line_ids": [
                    (0, 0, {"field_id": field_name.id, "is_required": True}),
                ],
            }
        )
        wiz = self._new_wizard(tpl)
        wb = self._attachment_from_action(wiz.action_export_data())
        self.assertIn("Data", wb.sheetnames)
