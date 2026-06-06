import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


def _join(items):
    return ", ".join(str(i) for i in items)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - guarded by manifest external_dependencies
    Workbook = load_workbook = None

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_TRUE = {"1", "true", "yes", "y", "t", "x", "✓"}
_FALSE = {"0", "false", "no", "n", "f", "", "-"}


class HrUaeXlsxImportWizard(models.TransientModel):
    _name = "hr.uae.xlsx.import.wizard"
    _description = "Excel Import / Export"

    # Upper bound for the exact display_name fallback scan when resolving a
    # many2one by name. Keeps the round-trip reliable for small relation models
    # without ever scanning very large ones (use 'id:<id>' there instead).
    _M2O_SCAN_LIMIT = 5000

    template_id = fields.Many2one(
        "hr.uae.xlsx.template",
        string="Template",
        required=True,
        domain="[('active', '=', True)]",
    )
    model_name = fields.Char(related="template_id.model_name", readonly=True)
    allow_export = fields.Boolean(related="template_id.allow_export", readonly=True)
    allow_import = fields.Boolean(related="template_id.allow_import", readonly=True)

    import_file = fields.Binary(string="File to Import")
    import_filename = fields.Char(string="File Name")

    result_html = fields.Html(string="Result", readonly=True, sanitize=False)
    create_count = fields.Integer(readonly=True)
    update_count = fields.Integer(readonly=True)
    error_count = fields.Integer(readonly=True)
    has_result = fields.Boolean(readonly=True)
    error_file = fields.Binary(string="Error Report", readonly=True)
    error_filename = fields.Char(readonly=True)

    # ===================================================================
    # small helpers
    # ===================================================================

    @staticmethod
    def _norm(value):
        return "" if value is None else str(value).strip()

    @classmethod
    def _is_empty(cls, value):
        return value is None or (isinstance(value, str) and value.strip() == "")

    def _check_openpyxl(self):
        if Workbook is None:
            raise UserError(
                _("The Python library 'openpyxl' is required for Excel import/export.")
            )

    # ===================================================================
    # value conversion (import)
    # ===================================================================

    def _to_number(self, raw):
        if isinstance(raw, bool):
            return 1 if raw else 0
        if isinstance(raw, (int, float)):
            return raw
        return float(str(raw).strip())

    def _to_bool(self, raw):
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        token = str(raw).strip().lower()
        if token in _TRUE:
            return True
        if token in _FALSE:
            return False
        raise ValidationError(_("'%s' is not a valid Yes/No value.") % raw)

    def _to_date(self, raw):
        if hasattr(raw, "date") and not isinstance(raw, str):
            return raw.date() if hasattr(raw, "hour") else raw
        value = fields.Date.to_date(str(raw).strip())
        if not value:
            raise ValidationError(_("'%s' is not a valid date (use YYYY-MM-DD).") % raw)
        return value

    def _to_datetime(self, raw):
        if not isinstance(raw, str) and hasattr(raw, "year"):
            return fields.Datetime.to_datetime(raw)
        value = fields.Datetime.to_datetime(str(raw).strip())
        if not value:
            raise ValidationError(_("'%s' is not a valid date/time.") % raw)
        return value

    def _resolve_selection(self, model, line, raw):
        field_obj = model._fields[line.name]
        options = field_obj._description_selection(self.env)
        raw_s = str(raw).strip()
        values = {str(v): v for v, _label in options}
        if raw_s in values:
            return values[raw_s]
        low = raw_s.lower()
        by_label = {str(label).strip().lower(): v for v, label in options}
        if low in by_label:
            return by_label[low]
        by_value = {str(v).lower(): v for v, _label in options}
        if low in by_value:
            return by_value[low]
        allowed = _join(["%s (%s)" % (label, v) for v, label in options])
        raise ValidationError(
            _("'%(val)s' is not allowed. Use one of: %(allowed)s.", val=raw_s, allowed=allowed)
        )

    def _resolve_one_m2o(self, comodel_name, raw):
        raw_s = str(raw).strip()
        if not raw_s:
            return False
        comodel = self.env[comodel_name]
        low = raw_s.lower()
        if low.startswith("id:") or low.startswith(".id:"):
            ref = raw_s.split(":", 1)[1].strip()
            try:
                rid = int(ref)
            except ValueError:
                raise ValidationError(_("'%s' is not a valid id.") % raw_s)
            rec = comodel.browse(rid)
            if not rec.exists():
                raise ValidationError(
                    _("%(model)s with id %(id)s does not exist.", model=comodel_name, id=rid)
                )
            return rec.id
        if low.startswith("xmlid:"):
            xmlid = raw_s.split(":", 1)[1].strip()
            rec = self.env.ref(xmlid, raise_if_not_found=False)
            if not rec:
                raise ValidationError(_("External id '%s' was not found.") % xmlid)
            return rec.id
        matches = comodel.name_search(raw_s, operator="=", limit=2)
        if len(matches) == 1:
            return matches[0][0]
        if len(matches) > 1:
            raise ValidationError(
                _(
                    "'%(val)s' matches several %(model)s records. Use 'id:<id>' "
                    "to be explicit.",
                    val=raw_s,
                    model=comodel_name,
                )
            )
        # Fallback: exact display_name match. This makes the export -> edit ->
        # re-import round-trip reliable for relations whose display_name is not
        # plainly searchable by name (code-prefixed analytic accounts, etc.).
        # Bounded so it never scans huge models like res.partner.
        scan = comodel.with_context(active_test=False)
        if scan.search_count([]) <= self._M2O_SCAN_LIMIT:
            candidates = scan.search([]).filtered(
                lambda r: r.display_name == raw_s
            )
            if len(candidates) == 1:
                return candidates.id
            if len(candidates) > 1:
                raise ValidationError(
                    _(
                        "'%(val)s' matches several %(model)s records. Use "
                        "'id:<id>' to be explicit.",
                        val=raw_s,
                        model=comodel_name,
                    )
                )
        raise ValidationError(
            _("No %(model)s named '%(val)s'.", model=comodel_name, val=raw_s)
        )

    def _resolve_x2many(self, comodel_name, raw):
        tokens = [t.strip() for t in str(raw).replace("\n", ",").split(",") if t.strip()]
        ids = [self._resolve_one_m2o(comodel_name, token) for token in tokens]
        return [(6, 0, ids)]

    def _cell_to_value(self, model, line, raw):
        ttype = line.field_type
        if ttype in ("char", "text", "html"):
            return str(raw).strip()
        if ttype == "integer":
            return int(self._to_number(raw))
        if ttype in ("float", "monetary"):
            return float(self._to_number(raw))
        if ttype == "boolean":
            return self._to_bool(raw)
        if ttype == "date":
            return self._to_date(raw)
        if ttype == "datetime":
            return self._to_datetime(raw)
        if ttype == "selection":
            return self._resolve_selection(model, line, raw)
        if ttype == "many2one":
            return self._resolve_one_m2o(line.relation, raw)
        if ttype in ("many2many", "one2many"):
            return self._resolve_x2many(line.relation, raw)
        raise ValidationError(_("Field type '%s' is not supported for import.") % ttype)

    # ===================================================================
    # export
    # ===================================================================

    def _format_export(self, model, line, record):
        ttype = line.field_type
        value = record[line.name]
        if ttype == "many2one":
            return value.display_name or "" if value else ""
        if ttype in ("many2many", "one2many"):
            return ", ".join(value.mapped("display_name"))
        if ttype == "boolean":
            return "TRUE" if value else "FALSE"
        if ttype == "selection":
            options = dict(model._fields[line.name]._description_selection(self.env))
            return options.get(value, value or "") if value not in (False, None) else ""
        if ttype == "date":
            return fields.Date.to_string(value) or ""
        if ttype == "datetime":
            return fields.Datetime.to_string(value) or ""
        if ttype in ("float", "monetary", "integer"):
            return value or 0
        if ttype == "binary":
            return "(binary — not exported)" if value else ""
        if value in (False, None):
            return ""
        return value

    def _build_instructions_sheet(self, workbook, template):
        ws = workbook.create_sheet("Instructions")
        meta = [
            ("Template", template.name),
            ("Model", template.model_name),
            ("Import allowed", _("Yes") if template.allow_import else _("No")),
            ("Create on import", _("Yes") if template.allow_create else _("No")),
            ("Update on import", _("Yes") if template.allow_update else _("No")),
        ]
        for key, val in meta:
            ws.append([key, val])
        ws.append([])
        if template.note:
            ws.append(["Notes", template.note])
            ws.append([])
        headers = [
            "Column",
            "Technical Field",
            "Type",
            "Importable",
            "Required",
            "Match Key",
            "Allowed values / Relation",
        ]
        ws.append(headers)
        header_row = ws.max_row
        model = template.get_model()
        for line in template.export_line_ids():
            detail = ""
            if line.field_type == "selection":
                options = model._fields[line.name]._description_selection(self.env)
                detail = _join(["%s (%s)" % (label, v) for v, label in options])
            elif line.field_type in ("many2one", "many2many", "one2many"):
                detail = _(
                    "Relation %(model)s — use a name, 'id:<id>' or 'xmlid:<ref>'",
                    model=line.relation,
                )
            importable = (
                line.for_import and not line.ref_only and template.allow_import
            )
            ws.append(
                [
                    line.label(),
                    line.name,
                    line.field_type,
                    _("Yes") if importable else _("No"),
                    _("Yes") if line.is_required else "",
                    _("Yes") if line.is_match_key else "",
                    detail,
                ]
            )
        ws.append([])
        ws.append(
            [
                "ID column",
                _(
                    "Keep the ID column to update existing records. Rows with an "
                    "empty ID create new records. Blank cells are left unchanged "
                    "on update."
                ),
            ]
        )
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["G"].width = 60

    def _build_workbook(self, template, records):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Data"
        export_lines = template.export_line_ids()
        labels = ["ID"] + [line.label() for line in export_lines]
        techs = ["id"] + [line.name for line in export_lines]
        ws.append(labels)
        ws.append(techs)
        model = template.get_model()
        for record in records:
            row = [record.id]
            for line in export_lines:
                row.append(self._format_export(model, line, record))
            ws.append(row)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.row_dimensions[2].hidden = True
        ws.freeze_panes = "A3"
        for idx, label in enumerate(labels, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(
                12, min(45, len(str(label)) + 6)
            )
        self._build_instructions_sheet(workbook, template)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _download(self, filename, content):
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "datas": base64.b64encode(content),
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": XLSX_MIME,
            }
        )
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%d?download=true" % attachment.id,
            "target": "self",
        }

    def _export(self, records):
        self.ensure_one()
        self._check_openpyxl()
        template = self.template_id
        if not template.allow_export:
            raise UserError(_("Export is disabled for this template."))
        content = self._build_workbook(template, records)
        filename = "%s_%s.xlsx" % (
            template.code,
            fields.Date.context_today(self).strftime("%Y%m%d"),
        )
        return self._download(filename, content)

    def action_export_data(self):
        self.ensure_one()
        records = self.template_id.get_model().search([])
        return self._export(records)

    def action_export_empty(self):
        self.ensure_one()
        return self._export(self.template_id.get_model().browse())

    # ===================================================================
    # import
    # ===================================================================

    def _column_mapping(self, rows, template):
        """Return (col_index -> tech field name) and the list of data rows."""
        import_lines = template.import_line_ids()
        tech_set = {line.name for line in import_lines}
        label_to_tech = {}
        for line in import_lines:
            label_to_tech[self._norm(line.label()).lower()] = line.name
            label_to_tech[self._norm(line.name).lower()] = line.name

        two_row = len(rows) >= 2 and self._norm(rows[1][0]).lower() == "id"
        col = {}
        if two_row:
            header, data, first_excel_row = rows[1], rows[2:], 3
            for idx, cell in enumerate(header):
                key = self._norm(cell).lower()
                if key == "id":
                    col[idx] = "id"
                elif self._norm(cell) in tech_set:
                    col[idx] = self._norm(cell)
        else:
            header, data, first_excel_row = rows[0], rows[1:], 2
            for idx, cell in enumerate(header):
                key = self._norm(cell).lower()
                if key in ("id", "id (do not edit)"):
                    col[idx] = "id"
                elif key in label_to_tech:
                    col[idx] = label_to_tech[key]
        mapped_fields = {v for v in col.values() if v != "id"}
        if not mapped_fields:
            raise UserError(
                _(
                    "No recognizable columns were found. Export a template first "
                    "and fill it in, keeping the header rows."
                )
            )
        return col, data, first_excel_row

    def _parse_file(self, template):
        self._check_openpyxl()
        try:
            raw = base64.b64decode(self.import_file)
            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(_("Could not read the Excel file: %s") % exc)
        sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.worksheets[0]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        workbook.close()
        # drop trailing fully-empty rows
        while rows and all(self._is_empty(c) for c in rows[-1]):
            rows.pop()
        if len(rows) < 2:
            raise UserError(_("The file has no data rows."))
        return self._column_mapping(rows, template)

    def _prepare_row(self, template, model, tech_to_line, col, raw_row):
        """Resolve a single row to (action, target_record, vals). Raises
        ValidationError with a friendly message on any problem."""
        id_val = None
        converted = {}
        for idx, fname in col.items():
            cell = raw_row[idx] if idx < len(raw_row) else None
            if fname == "id":
                id_val = cell
                continue
            if self._is_empty(cell):
                continue
            line = tech_to_line[fname]
            try:
                converted[fname] = self._cell_to_value(model, line, cell)
            except ValidationError as exc:
                raise ValidationError(
                    _("Column '%(c)s': %(e)s", c=line.label(), e=exc.args[0] if exc.args else exc)
                )
            except Exception as exc:
                raise ValidationError(_("Column '%(c)s': %(e)s", c=line.label(), e=exc))

        target = model.browse()
        if not self._is_empty(id_val):
            try:
                rid = int(self._to_number(id_val))
            except Exception:
                raise ValidationError(_("ID '%s' is not a valid record id.") % id_val)
            if not template.allow_update:
                raise ValidationError(_("This template does not allow updates."))
            target = model.browse(rid)
            if not target.exists():
                raise ValidationError(
                    _("Row references ID %s which does not exist.") % rid
                )
            action = "update"
        else:
            keys = [
                line for line in template.match_key_line_ids() if line.name in converted
            ]
            if keys:
                domain = [(line.name, "=", converted[line.name]) for line in keys]
                matches = model.with_context(active_test=False).search(domain)
                if len(matches) == 1:
                    if not template.allow_update:
                        raise ValidationError(_("This template does not allow updates."))
                    target, action = matches, "update"
                elif len(matches) > 1:
                    key_label = _join([line.label() for line in keys])
                    raise ValidationError(
                        _("Several records match on %s. Add the ID column to be explicit.")
                        % key_label
                    )
                else:
                    action = "create"
            else:
                action = "create"

        if action == "create":
            if not template.allow_create:
                raise ValidationError(
                    _("This template does not allow creating new records (no ID matched).")
                )
            for line in template.import_line_ids().filtered("is_required"):
                if line.name not in converted:
                    raise ValidationError(
                        _("Required column '%s' is empty.") % line.label()
                    )
        return action, target, converted

    def _run_import(self, commit):
        self.ensure_one()
        template = self.template_id
        if not template.allow_import or not template.import_line_ids():
            raise UserError(_("Import is disabled for this template."))
        if not self.import_file:
            raise UserError(_("Upload an Excel file first."))

        col, data, first_excel_row = self._parse_file(template)
        model = template.get_model()
        tech_to_line = {line.name: line for line in template.import_line_ids()}
        import_lines = template.import_line_ids()

        create_count = update_count = 0
        failures = []  # (excel_row, message, {label: raw})

        self.env.cr.execute("SAVEPOINT hr_uae_xlsx_import")
        for offset, raw_row in enumerate(data):
            excel_row = first_excel_row + offset
            if all(self._is_empty(c) for c in raw_row):
                continue
            try:
                with self.env.cr.savepoint():
                    action, target, vals = self._prepare_row(
                        template, model, tech_to_line, col, raw_row
                    )
                    if action == "create":
                        model.create(vals)
                        create_count += 1
                    else:
                        target.write(vals)
                        update_count += 1
                    self.env.flush_all()
            except Exception as exc:
                self.env.invalidate_all()
                message = exc.args[0] if isinstance(exc, (ValidationError, UserError)) and exc.args else str(exc)
                raw_map = {}
                for idx, fname in col.items():
                    if fname == "id":
                        raw_map["id"] = raw_row[idx] if idx < len(raw_row) else ""
                    elif fname in tech_to_line:
                        raw_map[fname] = raw_row[idx] if idx < len(raw_row) else ""
                failures.append((excel_row, message, raw_map))

        error_count = len(failures)
        persisted = commit and error_count == 0
        if persisted:
            self.env.cr.execute("RELEASE SAVEPOINT hr_uae_xlsx_import")
        else:
            self.env.cr.execute("ROLLBACK TO SAVEPOINT hr_uae_xlsx_import")

        self._store_result(
            template, import_lines, commit, persisted,
            create_count, update_count, failures,
        )
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "name": _("Excel Import / Export"),
        }

    def _store_result(self, template, import_lines, commit, persisted,
                       create_count, update_count, failures):
        error_count = len(failures)
        self.create_count = create_count
        self.update_count = update_count
        self.error_count = error_count
        self.has_result = True
        self.error_file = False
        self.error_filename = False

        if error_count:
            content = self._build_error_workbook(template, import_lines, failures)
            self.error_file = base64.b64encode(content)
            self.error_filename = "%s_errors.xlsx" % template.code

        if not commit:
            headline = _("Validation only — nothing was saved.")
            tone = "info"
        elif persisted:
            headline = _("Import complete. Changes were saved.")
            tone = "success"
        else:
            headline = _(
                "Import aborted because of errors — nothing was saved. Fix the "
                "rows below and import again."
            )
            tone = "danger"

        rows_html = ""
        for excel_row, message, _raw in failures[:50]:
            rows_html += (
                "<tr><td style='padding:2px 8px;'>%s</td>"
                "<td style='padding:2px 8px;color:#b00;'>%s</td></tr>"
                % (excel_row, _escape(message))
            )
        more = ""
        if error_count > 50:
            more = "<p>… and %s more (see the error report).</p>" % (error_count - 50)
        errors_block = ""
        if failures:
            errors_block = (
                "<h4>Errors</h4>"
                "<table style='border-collapse:collapse;'>"
                "<tr><th style='padding:2px 8px;'>Excel Row</th>"
                "<th style='padding:2px 8px;'>Problem</th></tr>"
                "%s</table>%s" % (rows_html, more)
            )
        self.result_html = (
            "<div class='alert alert-%s' role='alert'><b>%s</b></div>"
            "<ul><li>To create: <b>%s</b></li>"
            "<li>To update: <b>%s</b></li>"
            "<li>Errors: <b>%s</b></li></ul>%s"
            % (tone, _escape(headline), create_count, update_count, error_count, errors_block)
        )

    def _build_error_workbook(self, template, import_lines, failures):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Errors"
        labels = ["Excel Row", "Error", "ID"] + [line.label() for line in import_lines]
        ws.append(labels)
        for excel_row, message, raw_map in failures:
            row = [excel_row, message, raw_map.get("id", "")]
            for line in import_lines:
                value = raw_map.get(line.name, "")
                row.append("" if value is None else value)
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F8CBAD")
        ws.freeze_panes = "A2"
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def action_validate(self):
        return self._run_import(commit=False)

    def action_import(self):
        return self._run_import(commit=True)


def _escape(text):
    from markupsafe import escape

    return str(escape(text or ""))
